#!/usr/bin/python3

import openpyxl
from math import ceil
wb = openpyxl.load_workbook("student.xlsx")
ws=wb.active #활성 시트 불러오기 

min_row=2
total_list=[]
grade_list=[0,0]
Fstu = 0
for row in range(min_row,ws.max_row+1):
    midterm = ws.cell(column=3,row=row).value
    final = ws.cell(column=4,row=row).value
    homework = ws.cell(column=5,row=row).value
    attendance = ws.cell(column=6,row=row).value


    total = midterm * 0.3 + final * 0.35 + homework * 0.34 + attendance 
    if total<40:
        Fstu +=1
    
    total_list.append(total)

    ws.cell(column=7,row=row,value=total)
    
    
    
#정렬
# total_list2=sorted(total_list, key=lambda i:i[1],reverse=True) 
total_list2=sorted(total_list,reverse=True)

student=len(total_list2)

grade_list=[total_list2.index(t)+1 for t in total_list]

#3. A,B,C 비율 나누기

student2=student-Fstu
Astu= int(student2*0.3)
Bstu= int(student2*0.7)- Astu
Cstu = student2 -(Astu + Bstu)
# Astu = int(student2 * 0.3)
# Bstu = int(student2 * 0.4)
# Cstu = int(student2 * 0.3)
#4.플러스 제로 비율 나누기 
Aplus=int(Astu*0.5)
Bplus=int(Bstu*0.5)
Cplus=int(Cstu*0.5)
print(Astu,Bstu,Cstu,Aplus,Bplus,Cplus)
#학점 부여
for row in range(min_row,ws.max_row+1):
    if ws.cell(column=7, row=row).value <40:
        ws.cell(row=row, column=8).value = 'F'
    elif grade_list[row - min_row] <= Aplus:
        ws.cell(row=row,column=8).value='A+'
    elif grade_list[row - min_row] <=Astu:
        ws.cell(row=row, column=8).value = 'A0'
    elif grade_list[row - min_row] <=Astu+Bplus:
        ws.cell(row=row,column=8).value='B+'
    elif grade_list[row - min_row] <=Astu+Bstu:
        ws.cell(row=row, column=8).value = 'B0'
    elif grade_list[row - min_row] <=Astu+Bstu+Cplus:
        ws.cell(row=row, column=8).value = 'C+'
    else:
        ws.cell(row=row, column=8).value = 'C0'
# for i in range(Astu):
#     ws.cell(row=total_list2[i][0], column=8).value = 'A0'
# for i in range(Aplus):
#     ws.cell(row=total_list2[i][0], column=8).value = 'A+'
# for i in range(Astu,Astu+Bstu):
#     ws.cell(row=total_list2[i][0], column=8).value = 'B0'
# for i in range(Astu,Astu+Bplus):
#     ws.cell(row=total_list2[i][0], column=8).value = 'B+'
# for i in range(Astu+Bstu,Astu+Bstu+Cstu):
#     ws.cell(row=total_list2[i][0], column=8).value = 'C0'
# for i in range(Astu+Bstu,Astu+Bstu+Cplus):
#     ws.cell(row=total_list2[i][0], column=8).value = 'C+'
    
# for i in range(student2, student):
#     ws.cell(row=total_list2[i][0], column=8).value = 'F'        

    
wb.save('student.xlsx')
wb.close()